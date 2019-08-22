import openpyxl


def create_workbook_no_paid_list(no_paid_list):
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('미결제명단')

    sheet.append(['이름', '이메일주소', '학번', '결제여부', '결제일'])

    for row in no_paid_list:
        sheet.append(row)

    wb.save('D:\\tyn_dev\\workspace_pycham\\task-automation\\excel_files\\미결제명단.xlsx')


def main():
    wb = openpyxl.load_workbook('D:\\tyn_dev\\workspace_pycham\\task-automation\\excel_files\\수강생_결제정보.xlsx')
    sheet = wb['결제정보']

    no_paid_list = []

    for row_index in range(2, sheet.max_row+1):
        if sheet.cell(row_index, 4).value == '미결제':
            print(row_index)
            # 셀의 정보가 삽입된다.
            no_paid_list.append([
                sheet.cell(row_index, 1).value,
                sheet.cell(row_index, 2).value,
                sheet.cell(row_index, 3).value,
                sheet.cell(row_index, 4).value,
                sheet.cell(row_index, 5).value
            ])

    print(no_paid_list)
    create_workbook_no_paid_list(no_paid_list)

    print("Considerate Done!")

main()