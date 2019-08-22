'''
D:\tyn_dev\workspace_pycham\task-automation\venv\Scripts\python.exe "D:\tyn_dev\ide\PyCharm Community Edition 2019.2\helpers\pydev\pydevconsole.py" --mode=client --port=61314
import sys; print('Python %s on %s' % (sys.version, sys.platform))
sys.path.extend(['D:\\tyn_dev\\workspace_pycham\\task-automation', 'D:/tyn_dev/workspace_pycham/task-automation'])
PyDev console: starting.
Python 3.7.4 (tags/v3.7.4:e09359112e, Jul  8 2019, 20:34:20) [MSC v.1916 64 bit (AMD64)] on win32

pip install openpyxl

import openpyxl as excel

excel.Workbook()
<openpyxl.workbook.workbook.Workbook object at 0x000002633C324E08>
wb = excel.Workbook()
wb.get_sheet_names()
<input>:1: DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
['Sheet']
wb.get_sheet_names()
['Sheet']
wb.sheetnames
['Sheet']
wb.active
<Worksheet "Sheet">
wb.get_active_sheet()
<input>:1: DeprecationWarning: Call to deprecated function get_active_sheet (Use the .active property).
<Worksheet "Sheet">
sheet = wb.get_sheet_by_name('Sheet')
<input>:1: DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
sheet = wb['Sheet']
sheet['A1'].value
sheet['A1'].value = 'hello world'
sheet['A1'].value
'hello world'
sheet['B1'] = 50000
sheet['B1']
<Cell 'Sheet'.B1>
sheet['B1'].value
50000
sheet.cell(3, 2, '안녕하세요!')
<Cell 'Sheet'.B3>
sheet['B3']
<Cell 'Sheet'.B3>
sheet['B3'].value
'안녕하세요!'
wb.save('excel-python-domo.xlsx')
wb2 = excel.load_workbook('excel-python-domo.xlsx')
wb2.active
<Worksheet "Sheet">
sheet2 = wb2.active
sheet2.cell(1,1)
<Cell 'Sheet'.A1>
sheet2.cell(1,1).value
'hello world'
sheet2.title
'Sheet'
sheet2.title = '바꾼 시트 이름'
sheet2.title
'바꾼 시트 이름'
'''